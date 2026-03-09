from __future__ import annotations

import csv
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

SKIP_PATTERNS = ("Consulta Personalizada", "Looker", "Dados Estáticos")


@dataclass
class ColumnLayout:
    col_endereco_tabela: int = 4
    col_data_ref_base: int = 8
    col_data_ref_painel: int = 11
    csv_delimiter: str = ";"
    csv_encoding: str = "utf-8-sig"


@dataclass
class CsvRow:
    indice: int
    dados: list[str]
    endereco_tabela: str = ""
    skip: bool = False
    motivo_skip: str = ""

    def init_from_layout(self, layout: ColumnLayout) -> None:
        if len(self.dados) > layout.col_endereco_tabela:
            self.endereco_tabela = self.dados[layout.col_endereco_tabela].strip()
        if not self.endereco_tabela or self._is_vazia():
            self.skip = True
            self.motivo_skip = "linha vazia" if self._is_vazia() else "sem endereco"
        elif any(p in self.endereco_tabela for p in SKIP_PATTERNS):
            self.skip = True
            self.motivo_skip = "consulta personalizada"

    def _is_vazia(self) -> bool:
        return all(c.strip() == "" for c in self.dados)


def _criar_rows(data: list[list[str]], layout: ColumnLayout) -> list[CsvRow]:
    rows: list[CsvRow] = []
    for i, raw in enumerate(data, start=1):
        row = CsvRow(indice=i, dados=raw)
        row.init_from_layout(layout)
        rows.append(row)
    return rows


@dataclass
class CsvData:
    header: list[str]
    rows: list[CsvRow]
    fonte: str = ""

    @property
    def tabelas_unicas(self) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for row in self.rows:
            if not row.skip and row.endereco_tabela and row.endereco_tabela not in seen:
                seen.add(row.endereco_tabela)
                result.append(row.endereco_tabela)
        return result

    @property
    def stats(self) -> dict[str, int]:
        total = len(self.rows)
        vazias = sum(1 for r in self.rows if r.motivo_skip == "linha vazia")
        skip = sum(1 for r in self.rows if r.skip and r.motivo_skip != "linha vazia")
        validas = total - vazias - skip
        return {
            "total": total,
            "vazias": vazias,
            "skip": skip,
            "validas": validas,
            "tabelas_unicas": len(self.tabelas_unicas),
        }


def ler_csv(caminho: Path, layout: ColumnLayout | None = None) -> CsvData:
    layout = layout or ColumnLayout()
    logger.info("Lendo CSV: %s", caminho)
    with open(caminho, "r", encoding=layout.csv_encoding, newline="") as f:
        reader = csv.reader(f, delimiter=layout.csv_delimiter)
        header = next(reader)
        data = [raw for raw in reader]
    rows = _criar_rows(data, layout)
    logger.info("CSV lido: %d linhas", len(rows))
    return CsvData(header=header, rows=rows, fonte=str(caminho))


def ler_xlsx(
    caminho: Path,
    aba: str | None = None,
    layout: ColumnLayout | None = None,
) -> CsvData:
    layout = layout or ColumnLayout()
    logger.info("Lendo XLSX: %s (aba=%s)", caminho, aba)
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb[aba] if aba else wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(c) if c is not None else "" for c in row])
    wb.close()

    if not data:
        return CsvData(header=[], rows=[], fonte=str(caminho))

    header = data[0]
    rows = _criar_rows(data[1:], layout)
    logger.info("XLSX lido: %d linhas", len(rows))
    return CsvData(header=header, rows=rows, fonte=str(caminho))


def ler_arquivo(
    caminho: Path,
    aba: str | None = None,
    layout: ColumnLayout | None = None,
) -> CsvData:
    sufixo = caminho.suffix.lower()
    if sufixo == ".xlsx":
        return ler_xlsx(caminho, aba, layout)
    return ler_csv(caminho, layout)


def listar_abas_xlsx(caminho: Path) -> list[str]:
    from openpyxl import load_workbook

    wb = load_workbook(caminho, read_only=True)
    nomes = wb.sheetnames
    wb.close()
    return nomes


def salvar_csv(csv_data: CsvData, caminho_saida: Path, layout: ColumnLayout | None = None) -> Path:
    layout = layout or ColumnLayout()
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_base = Path(csv_data.fonte).stem if csv_data.fonte else "output"
    arquivo = caminho_saida / f"{nome_base}_{timestamp}.csv"

    with open(arquivo, "w", encoding=layout.csv_encoding, newline="") as f:
        writer = csv.writer(f, delimiter=layout.csv_delimiter)
        writer.writerow(csv_data.header)
        for row in csv_data.rows:
            writer.writerow(row.dados)

    logger.info("CSV salvo: %s", arquivo)
    return arquivo


def atualizar_data_linha(
    row: CsvRow,
    col_index: int,
    novo_valor: str,
) -> str:
    while len(row.dados) <= col_index:
        row.dados.append("")
    antigo = row.dados[col_index]
    row.dados[col_index] = novo_valor
    return antigo
